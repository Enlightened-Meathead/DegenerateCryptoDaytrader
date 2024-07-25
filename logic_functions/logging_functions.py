## Functions that take information from buy and sell orders then log it to a spreadsheet
import os
from openpyxl import Workbook, load_workbook

from resources import config
from datetime import datetime

workbook_file_path = config.trade_log_path


def check_workbook_existence(filepath):
    if not os.path.exists(filepath):
        while True:
            user_response = input(
                f"The excel workbook filepath in your config does not exist, would you like to create "
                f"one at {workbook_file_path}? yes/no: ").strip().lower()
            if user_response == 'yes':
                create_initial_workbook()
                print('New workbook created!')
                break
            elif user_response == 'no':
                print("Aborting logging the trade. If your trade log already exists, please enter an absolute "
                      "filepath in the program config file. If the log doesn't exist, enter the filepath you want it "
                      "to exist in the config file and run "
                      "this program with the --create_workbook option to prompt this workbook creation menu again.")
                return False
            else:
                print("Enter 'yes' or 'no'")


# If you manually create a workbook in Excel or in my case LibreOffice Calc beforehand, you'll have issues with openpyxl
def create_initial_workbook():
    workbook = Workbook()
    sheet = workbook.active
    headers = ["Order Number", "Date and Time", "Order Type", "Asset", "Bought Price", "Asset Amount Bought",
               "$ Amount Bought", "Sold Price", "Amount of Asset Sold", "$ Amount Sold", "Profit/Loss %",
               "Profit/Loss $ Amount", "$ Total Losses", "$ Total Profits", "$ Total Net"]
    sheet.append(headers)
    workbook.save(workbook_file_path)


# Calculate the profits, losses, and total net gains in the excel document
def calculate_totals():
    workbook = load_workbook(workbook_file_path)
    sheet = workbook.active
    total_losses = 0
    total_profits = 0
    # Get the profit or loss of each row and calculate the total losses, total profits, and net
    incorrect_rows = 0
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=12, max_col=12, values_only=True):
        try:
            profit_loss = row[0]
            if profit_loss < 0:
                total_losses += profit_loss
            elif profit_loss > 0:
                total_profits += profit_loss
        except Exception as e:
            incorrect_rows = 1
            # This function will always find the wrong data type in the first row, so this insures only real errors
            # get reported
            if incorrect_rows > 1:
                print(f"Incorrect data type in calculate_totals from logging_functions: {e}")

    total_net = total_profits + total_losses

    # Write the totals in the last three columns
    sheet.cell(row=2, column=13, value=round(total_losses, 2))
    sheet.cell(row=2, column=14, value=round(total_profits, 2))
    sheet.cell(row=2, column=15, value=round(total_net, 2))

    try:
        workbook.save(workbook_file_path)
        workbook.close()
    except Exception as e:
        print(f"Error in log_trade trying to save to workbook: {e}")


# Log the trade and append it to the trade_log spreadsheet
def log_trade(*args):
    # Load in the workbook and select the main sheet then find the current free row
    try:
        workbook = load_workbook(workbook_file_path)
        sheet = workbook.active
        next_row = sheet.max_row + 1
        order_number = sheet.max_row
        # Create a list of values to be appended from arguments passed to the function
        values_to_append = list(args)
        values_to_append = [order_number] + values_to_append
        # Add each value to their cells
        for column, value in enumerate(values_to_append, start=1):
            try:
                value = float(value)
            except:
                pass
            sheet.cell(row=next_row, column=column, value=value)
        # Save the changes to the spreadsheet
        try:
            workbook.save(workbook_file_path)
            workbook.close()
            print(f"Order appended to {workbook_file_path}")
        except Exception as e:
            print(f"Error in log_trade trying to save to workbook: {e}")
    except FileNotFoundError:
        print("Workbook not found! Error in log_trade function")


# Sort the one-liner values so click doesn't read one that would cause the menu to start before reading all inputs
def sort_one_liner(selected_user_options):
    # The order the options should be interpreted by the click command parser from last to first
    keys_list = [
        "menu",
        "bot_type",
        "asset",
        "capital",
        "initial_capital",
        "start_type",
        "log_trade",
        "sell_order_type",
        "asset_bought_price",
        "percent_loss_limit",
        "profit_loss_function",
        "buy_order_type",
        "basic_buy_price",
        "rsi_buy_number",
        "rsi_drop_limit",
        "rsi_wait_period",
        "basic_sell_profit",
        "minimum_ladder_profit",
        "ladder_step_gain",
        "ladder_step_loss",
        "ladder_timer_duration",
        "ladder_step_sensitivity",
        "ladder_timer_sensitivity",
        "swing_trade_skim",
        "history",
        "message_history"
    ]
    # I wrote these after learning about comprehensions if you can't tell ;)
    sorted_option_keys = sorted(selected_user_options, key=lambda k: keys_list.index(k), reverse=True)
    sorted_user_options = {k: selected_user_options[k] for k in sorted_option_keys}
    return sorted_user_options


# Repeat the users command options and save it to the history file, so they can copy and paste it and not have to
# re-enter everything if they made a mistake or want to do a similar trade again in the future
def repeat_one_liner(selected_user_options):
    program_alias = config.repeat_one_liner_alias
    # Take the user options and sort them into the proper order for click
    sorted_user_options = sort_one_liner(selected_user_options)
    print("\n===========================\nCurrent options one-liner:\n===========================")
    # Add the --option value strings to a list then join the list
    command_option_list = [f"--{option} {value}" for option, value in sorted_user_options.items()]
    command_option_string = ' '.join([string for string in command_option_list])
    current_one_liner = f'python3 {program_alias}.py {command_option_string}'
    print(current_one_liner)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Only log the one-liner if this function is called from another file
    if __name__ != "__main__":
        with open('./resources/dcd_command_history.txt', 'a') as history:
            history.write(f'\n{timestamp} : {current_one_liner}\n')


def read_command_history(ctx, param, value):
    if value == "view":
        with open('resources/dcd_command_history.txt', 'r') as history:
            print(history.read())
            history.close()
            exit()
    elif value == "clear":
        # Open the file with write, which clears the contents of the file
        with open('./resources/dcd_command_history.txt', 'w'):
            print("Command history cleared!")
            exit()


# Log message notifications to a history file in case the user misses the notification or email
def log_message(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open('./resources/dcd_message_history.txt', 'a') as history:
        history.write(f"\n{timestamp} : {message}\n")


def read_message_history(ctx, param, value):
    if value == "view":
        with open('resources/dcd_message_history.txt', 'r') as history:
            print(history.read())
            exit()
    elif value == "clear":
        # Open the file with write, which clears the contents of the file
        with open('./resources/dcd_message_history.txt', 'w'):
            print("Message history cleared!")
            exit()


if __name__ == "__main__":
    pass
